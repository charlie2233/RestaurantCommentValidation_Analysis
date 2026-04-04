from __future__ import annotations

import filecmp
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from qsr_audit.ingest.workbook import load_workbook_sheets
from qsr_audit.normalize.parsing import canonicalize_brand_name, parse_fte_range, parse_margin_range
from typer.testing import CliRunner

from qsr_audit.cli import app

TOP30_SHEET = "QSR Top30 核心数据"
STRATEGY_SHEET = "AI策略与落地效果"
NOTES_SHEET = "数据说明与来源"

TOP30_HEADERS = [
    "排名",
    "品牌",
    "品类",
    "美国门店数\n(2024)",
    "全系统营收\n($B, 2024)",
    "店均AUV\n($K)",
    "店均日等效FTE\n(估算)",
    "门店利润率\n(估算)",
    "央厨/供应链模式",
    "所有制模式",
]

STRATEGY_HEADERS = [
    "品牌",
    "AI/技术策略方向",
    "关键举措",
    "部署规模",
    "落地效果/数据",
    "当前状态(2026Q1)",
]

NOTES_HEADERS = ["字段", "说明"]


def build_fixture_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = TOP30_SHEET
    ws.append(TOP30_HEADERS)
    ws.append(
        [
            1,
            "  mcdonald's  ",
            "汉堡",
            19502,
            53.5,
            4001,
            "25-35",
            "18-20%",
            "肉饼由区域供应商冷冻预成型；门店用烤炉+油炸锅完成",
            "95%加盟",
        ]
    )
    ws.append(
        [
            2,
            "dunkin",
            "咖啡/甜甜圈",
            9509,
            11.4,
            1200,
            "10-15",
            "15-18%",
            "甜甜圈由央厨或区域供应商预制冷冻→门店烘烤/装饰；饮品门店现场制作",
            "100%加盟",
        ]
    )
    ws.append(
        [
            3,
            "in-n-out",
            "汉堡",
            410,
            1.25,
            3130,
            "20-25",
            "20-24%",
            "自有加工厂→冷链直送；薯条门店从整颗土豆现切现炸",
            "100%直营",
        ]
    )

    ws = wb.create_sheet(STRATEGY_SHEET)
    ws.append(STRATEGY_HEADERS)
    ws.append(
        [
            "mcdonald's",
            "语音AI点单→转向后台AI",
            "①2021-2024与IBM测试Drive-thru AOT语音点单\n②与Google合作GenAI\n③预测性设备维护",
            "AOT: 100家→已关停",
            "AOT准确率不达标；Google合作转向后台运营优化。",
            "AOT终止；聚焦后台AI",
        ]
    )
    ws.append(
        [
            "DUNKIN",
            "门店自动化 + 数字化",
            "①数字菜单\n②忠诚度App\n③排班优化",
            "有限",
            "移动端订单增长。",
            "追赶期",
        ]
    )
    ws.append(
        [
            "In N Out",
            "低技术、高流程",
            "①极简菜单\n②强流程标准化",
            "门店全系统",
            "保持高AUV与高一致性。",
            "保持低技术策略",
        ]
    )

    ws = wb.create_sheet(NOTES_SHEET)
    ws.append(NOTES_HEADERS)
    ws.append(["美国门店数", "2024年底数据，来源QSR 50 2025 Report / 各品牌10-K年报 / Statista"])
    ws.append(["全系统营收", "美国系统营收（含加盟商），2024年。"])
    ws.append(["店均AUV", "Average Unit Volume，全系统美国平均值。"])
    ws.append(
        [
            "店均日等效FTE",
            "基于公开数据估算的每门店日均等效全职人数（8小时=1 FTE）。仅为估算区间。",
        ]
    )
    ws.append(["门店利润率", "Restaurant-Level Operating Margin，不含总部费用/折旧/租金。"])
    ws.append(["央厨/供应链模式", "描述食品从原料到成品的处理链路。"])
    ws.append(["AI策略", "基于公开新闻报道、品牌官方公告、投资者电话会议、行业分析整理。"])
    ws.append([None, None])
    ws.append(["关键发现", None])
    ws.append(
        ["1", "AUV排名: Chick-fil-A > Raising Cane's > McDonald's > Shake Shack > Whataburger"]
    )
    ws.append(["2", "AI策略光谱: 全自动化 → 协作机器人 → 调度AI → AI语音 → 纯人工"])
    ws.append(["3", "语音AI Drive-thru: McDonald's失败退出，Taco Bell遭挫重新评估。"])
    ws.append(["4", "物理自动化: Sweetgreen 和 Chipotle 是推进厨房机器人的代表。"])
    ws.append(["5", "反技术路线也能成功。"])
    ws.append(["6", "门店现做程度与AUV高度正相关。"])
    ws.append([None, None])
    ws.append(["数据收集日期", "2026年4月3日"])
    ws.append(
        ["主要来源", "QSR Magazine QSR 50 (2025), Technomic Top 500, Restaurant Business, NRN"]
    )
    ws.append(["用途", "AtomBite.ai 竞争格局分析与投资者材料参考"])

    wb.save(path)
    return path


@pytest.fixture()
def fixture_workbook_path(tmp_path: Path) -> Path:
    return build_fixture_workbook(tmp_path / "source" / "fixture_workbook.xlsx")


def test_parse_fte_range_handles_ranges():
    assert parse_fte_range("6-8") == (6, 8, 7.0)
    assert parse_fte_range("20-25") == (20, 25, 22.5)


def test_parse_margin_range_handles_percentage_ranges():
    assert parse_margin_range("8-12%") == (8, 12, 10.0)
    assert parse_margin_range("18-22%") == (18, 22, 20.0)


@pytest.mark.parametrize(
    ("raw_name", "expected"),
    [
        ("  mcdonald's  ", "McDonald's"),
        ("dunkin", "Dunkin'"),
        ("In N Out", "In-N-Out"),
    ],
)
def test_brand_canonicalization(raw_name: str, expected: str):
    assert canonicalize_brand_name(raw_name) == expected


def test_sheet_load_reads_all_three_sheets(fixture_workbook_path: Path):
    sheets = load_workbook_sheets(fixture_workbook_path)

    assert set(sheets) == {TOP30_SHEET, STRATEGY_SHEET, NOTES_SHEET}
    assert list(sheets[TOP30_SHEET].columns) == TOP30_HEADERS
    assert list(sheets[STRATEGY_SHEET].columns) == STRATEGY_HEADERS
    assert list(sheets[NOTES_SHEET].columns) == NOTES_HEADERS
    assert sheets[TOP30_SHEET].shape[0] == 3
    assert sheets[STRATEGY_SHEET].shape[0] == 3
    assert sheets[NOTES_SHEET].shape[0] == 19


def test_ingest_workbook_writes_bronze_and_silver_outputs(
    fixture_workbook_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    runner = CliRunner()
    monkeypatch.chdir(tmp_path)

    raw_dir = tmp_path / "data" / "raw"
    bronze_dir = tmp_path / "data" / "bronze"
    silver_dir = tmp_path / "data" / "silver"
    raw_dir.mkdir(parents=True, exist_ok=True)
    bronze_dir.mkdir(parents=True, exist_ok=True)
    silver_dir.mkdir(parents=True, exist_ok=True)

    input_path = raw_dir / fixture_workbook_path.name
    input_path.write_bytes(fixture_workbook_path.read_bytes())

    result = runner.invoke(app, ["ingest-workbook", "--input", str(input_path)])
    assert result.exit_code == 0, result.output

    copied_workbooks = [
        path for path in bronze_dir.rglob("*.xlsx") if filecmp.cmp(path, input_path, shallow=False)
    ]
    assert copied_workbooks, "expected the raw workbook copy to be preserved in bronze"

    assert len(list(bronze_dir.rglob("*.parquet"))) >= 3
    assert len(list(bronze_dir.rglob("*.csv"))) >= 3

    expected_silver_files = {
        "core_brand_metrics.parquet",
        "ai_strategy_registry.parquet",
        "data_notes.parquet",
        "key_findings.parquet",
    }
    assert expected_silver_files <= {path.name for path in silver_dir.glob("*.parquet")}

    core = pd.read_parquet(silver_dir / "core_brand_metrics.parquet")
    strategy = pd.read_parquet(silver_dir / "ai_strategy_registry.parquet")
    notes = pd.read_parquet(silver_dir / "data_notes.parquet")
    findings = pd.read_parquet(silver_dir / "key_findings.parquet")

    assert core["brand_name"].tolist() == ["McDonald's", "Dunkin'", "In-N-Out"]
    assert core["fte_min"].tolist() == [25, 10, 20]
    assert core["fte_max"].tolist() == [35, 15, 25]
    assert core["fte_mid"].tolist() == [30.0, 12.5, 22.5]
    assert core["margin_min_pct"].tolist() == [18, 15, 20]
    assert core["margin_max_pct"].tolist() == [20, 18, 24]
    assert core["margin_mid_pct"].tolist() == [19.0, 16.5, 22.0]
    assert core["source_sheet"].tolist() == [TOP30_SHEET, TOP30_SHEET, TOP30_SHEET]
    assert core["row_number"].tolist() == [2, 3, 4]

    assert strategy["brand_name"].tolist() == ["McDonald's", "Dunkin'", "In-N-Out"]
    assert strategy["source_sheet"].tolist() == [STRATEGY_SHEET, STRATEGY_SHEET, STRATEGY_SHEET]
    assert strategy["row_number"].tolist() == [2, 3, 4]
    assert any("\n" in text for text in strategy["key_initiatives"].tolist())

    assert notes["source_sheet"].tolist() == [NOTES_SHEET] * 7
    assert notes["row_number"].tolist() == [2, 3, 4, 5, 6, 7, 8]

    assert findings["source_sheet"].tolist() == [NOTES_SHEET] * 6
    assert findings["row_number"].tolist() == [11, 12, 13, 14, 15, 16]
